import os
from openpyxl import load_workbook
import re
from PrimaryParser.chinese import ChinesePrimary
from PrimaryParser.english import EnglishPrimary
from PrimaryParser.math import MathPrimary
from PrimaryParser.physics import PhysicsPrimary
from PrimaryParser.chemistry import ChemistryPrimary
from PrimaryParser.biology import BiologyPrimary
from PrimaryParser.history import HistoryPrimary
from PrimaryParser.politics import PoliticsPrimary
from PrimaryParser.geo import GeoPrimary
import asyncio
import argparse


def run_parser(max_threads=256):
    rid = 0
    exam_id_pattern = re.compile(r'rid=(\d+)')
    school_name_pattern = re.compile(r'^(\D+)\d+')

    home_path = os.environ['HOME']
    source_file_path = home_path + '/GaoKao/'
    gen_file_path = home_path + '/generate/'
    if not os.path.exists(gen_file_path):
        os.mkdir(gen_file_path)
    json_root_path = gen_file_path + 'json/'
    error_root_path = gen_file_path + 'error/'
    if not os.path.exists(json_root_path):
        os.mkdir(json_root_path)
    if not os.path.exists(error_root_path):
        os.mkdir(error_root_path)
    excel_root_path = source_file_path + 'excel/'
    if not os.path.exists(excel_root_path):
        raise Exception('Excel Path ' + excel_root_path + " does not exist")
    if not os.path.exists('%simg/' % gen_file_path):
        os.mkdir('%simg/' % gen_file_path)
    excel_list = next(os.walk(excel_root_path))[2]
    for excel_name in excel_list:
        if excel_name == '.DS_Store':
            continue
        print('parsing %s' % excel_name)
        total_count = 0
        problem_count = 0
        excel_path = excel_root_path + excel_name
        work_book = load_workbook(excel_path)
        work_sheet = work_book.active
        exam_root_path = source_file_path + 'source_html/' + excel_name.split('.')[0] + '/'
        img_root_path = '%simg/%s/' % (gen_file_path, excel_name.split('.')[0])
        if not os.path.exists(img_root_path):
            os.mkdir(img_root_path)
        img_list_path = '%simage_list.txt' % img_root_path
        # if os.path.exists(img_list_path):
        #     os.remove(img_list_path)
        arg_list = []
        error_type_path = ""
        if '2021' in excel_path:
            x = 0
        for row in work_sheet.iter_rows(min_row=2):
            name = row[0].value
            url = row[1].value
            year = row[2].value
            subject = row[3].value
            grade = row[4].value
            if len(row) >= 6:
                exam_id = row[5].value
            else:
                exam_id = exam_id_pattern.search(url).group(1)
            exam_path = exam_root_path + exam_id + '/'
            if not os.path.exists(exam_path):
                raise Exception('Error: exam path ' + exam_path + ' does not exists')
            # create exam json folder
            json_type_path = '%s%s/' % (json_root_path, excel_name.split('.')[0])
            error_type_path = '%s%s/' % (error_root_path, excel_name.split('.')[0])
            if not os.path.exists(json_type_path):
                os.mkdir(json_type_path)
            if not os.path.exists(error_type_path):
                os.mkdir(error_type_path)
            exam_json_path = '%s%s/' % (json_type_path, exam_id)
            error_file_path = '%s%s/' % (error_type_path, exam_id)
            if not os.path.exists(exam_json_path):
                os.mkdir(exam_json_path)
            if not os.path.exists(error_file_path):
                os.mkdir(error_file_path)
            html_list = next(os.walk(exam_path))[2]
            question_html_list = []
            answer_html_list = []
            for html_name in html_list:
                if html_name == "%s.html" % exam_id or html_name.endswith('.txt') or html_name == '.DS_Store':
                    continue
                if html_name.split('.')[0].split('_')[-1] in ['answer', 'Answer']:
                    answer_html_list.append(html_name)
                else:
                    question_html_list.append(html_name)
            if len(answer_html_list) != len(question_html_list):
                print(answer_html_list)
                print(len(answer_html_list))
                print(question_html_list)
                print(len(question_html_list))
                raise AssertionError
            for i in range(len(question_html_list)):
                paths = {
                    'problem_path': '%s%s' % (exam_path, question_html_list[i]),
                    'answer_path': '%s%s_Answer.html' % (exam_path, question_html_list[i].split('.')[0]),
                    'output_path': '%s%s.json' % (exam_json_path, question_html_list[i].split('.')[0]),
                    'error_path': '%s%s.json' % (error_file_path, question_html_list[i].split('.')[0]),
                    'image_path': img_root_path,
                    'image_list_path': img_list_path
                }
                school_name = school_name_pattern.search(name)
                if school_name is not None:
                    school_name = school_name.group(1)
                args = {
                    'subject': subject,
                    'grade': int(grade),
                    'source_link': url,
                    'year': year,
                    'test_name': name,
                    'school': school_name
                }
                arg_list.append((paths, args))
                total_count += 1

        print('Num Total: %d' % total_count)
        assert total_count == len(arg_list)
        if total_count == 0:
            continue
        num_group = total_count // max_threads + 1
        loop = asyncio.get_event_loop()
        processed_count = 0
        success_count = 0
        for i in range(num_group):
            tasks = []
            for j in range(min(max_threads, total_count - i * max_threads)):
                processed_count += 1
                current_path, current_args = arg_list[i * max_threads + j]
                subject = current_args['subject']
                if subject == 'chinese':
                    tasks.append(asyncio.ensure_future(ChinesePrimary(current_path, current_args).parse()))
                elif subject == 'english':
                    tasks.append(asyncio.ensure_future(EnglishPrimary(current_path, current_args).parse()))
                elif subject == 'math':
                    tasks.append(asyncio.ensure_future(MathPrimary(current_path, current_args).parse()))
                elif subject == 'physics':
                    tasks.append(asyncio.ensure_future(PhysicsPrimary(current_path, current_args).parse()))
                elif subject == 'chemistry':
                    tasks.append(asyncio.ensure_future(ChemistryPrimary(current_path, current_args).parse()))
                elif subject == 'biology':
                    tasks.append(asyncio.ensure_future(BiologyPrimary(current_path, current_args).parse()))
                elif subject == 'history':
                    tasks.append(asyncio.ensure_future(HistoryPrimary(current_path, current_args).parse()))
                elif subject == 'politics':
                    tasks.append(asyncio.ensure_future(PoliticsPrimary(current_path, current_args).parse()))
                elif subject == 'geo':
                    tasks.append(asyncio.ensure_future(GeoPrimary(current_path, current_args).parse()))
            future_result = loop.run_until_complete(asyncio.wait(tasks))
            result_list = list(future_result[0])
            for result in result_list:
                if result.result() is not None:
                    success_count += 1
                    problem_count += result.result()
            print('\rproc: %d, success: %d, rate: %.2f%%' %
                  (processed_count, success_count, success_count / processed_count * 100), end='')

        print('\nNum Problems: %d' % problem_count)

        # delete empty folder
        for dir, folder, file in os.walk(error_type_path):
            if folder == [] and file == []:
                os.rmdir(dir)


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('-t', '--thread', type=int, help='max threads')
    args = parser.parse_args()
    run_parser(args.thread)
