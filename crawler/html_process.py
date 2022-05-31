from bs4 import BeautifulSoup
import os
from openpyxl import load_workbook


def transfer(qhtml, ahtml, qhtml_save, ahtml_save):
    # q: question; a: answer
    qfile = open(qhtml, 'r')
    afile = open(ahtml, 'r')
    qsoup = BeautifulSoup(qfile.read(), 'lxml')
    asoup = BeautifulSoup(afile.read(), 'lxml')
    qfile.close()
    afile.close()
    analyse_nodes = qsoup.select('.view-analyse')
    for analyse_node in analyse_nodes:
        analyse_node.decompose()
    with open(qhtml_save, 'w') as f:
        f.write(str(qsoup))
    answer_nodes = asoup.select('.answer-item')
    with open(ahtml_save, 'w') as f:
        for answer_node in answer_nodes:
            f.write(str(answer_node))

origin_path = '/Users/flagerlee/GaoKao_origin/source_html'
new_path = '/Users/flagerlee/GaoKao/source_html'

excel_path = '/Users/flagerlee/GaoKao_origin/excel'

for name in os.listdir(excel_path):
    print('start process %s' % name)
    if name == '.DS_Store':
        continue
    if name.startswith('~$'):
        continue
    work_sheet = load_workbook('%s/%s' % (excel_path, name)).active
    tot_row = work_sheet.max_row - 1
    row_cnt = 1
    for row in work_sheet.iter_rows(min_row=2):
        year = row[2].value
        rid = row[5].value
        origin_html_path = '%s/%sGaoKao/%s' % (origin_path, year, rid)
        new_html_path = '%s/%sGaoKao/%s' % (new_path, year, rid)
        html_list = []
        for name in os.listdir(origin_html_path):
            if name == '.DS_Store':
                continue
            if name.endswith('.txt'):
                continue
            if name.split('.')[0].endswith('answer'):
                continue
            html_list.append(name.split('.')[0])
        for name in html_list:
            transfer(
                qhtml='%s/%s.html' % (origin_html_path, name),
                ahtml='%s/%s_answer.html' % (origin_html_path, name),
                qhtml_save='%s/%s.html' % (new_html_path, name),
                ahtml_save='%s/%s_answer.html' % (new_html_path, name)
            )
        print('processed: %.2f%%' % (row_cnt / tot_row * 100), end='\r')
        row_cnt += 1
    print('')