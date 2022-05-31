import requests
import argparse
import openpyxl
from bs4 import BeautifulSoup
import os
import re


subject_dict = {
    'chinese': '1',
    'math': '2',
    'history': '3',
    'biology': '4',
    'english': '5',
    'geo': '6',
    'politics': '7',
    'chemistry': '8',
    'physics': '9',
}
exam_type_dict = {
    '高考真题': '1',
    '高考模拟': '3',
    '高中联考': '2',
    '期中试卷': '4',
    '期末试卷': '5',
    '月考试卷': '6',
}


def run_spider(subject, year, grade, exam_type, excel_path, html_path):
    assert excel_path is not None
    assert html_path is not None
    assert html_path[-1] == '/'
    if subject is None:
        raise Exception('Missing argument "subject"')

    try:
        subject_id = subject_dict[subject]
    except KeyError as e:
        print('Key ' + subject + ' do not exists')
        return

    try:
        type_id = exam_type_dict[exam_type]
    except KeyError as e:
        print('Key ' + exam_type + ' do not exists')
        return

    root_url_prefix = 'http://5utk.ks5u.com/main.aspx?mod=paper&ac=st&op=list&page='
    root_url_suffix = '&rn=10&bankid=' + subject_id + '&Years=' + year + '&q2=0&q3=' + type_id + '&Grade=' + grade + \
                      '&lm=new'
    print(root_url_prefix + '1' + root_url_suffix)

    header = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 '
                            '(KHTML, like Gecko) Version/14.1.2 Safari/605.1.15'}

    # open new xls
    if not os.path.exists(excel_path):
        work_book = openpyxl.Workbook()
        work_sheet = work_book.active
        table_title = ['Name', 'Url', 'Year', 'Subject', 'Grade']
        for col in range(len(table_title)):
            work_sheet.cell(row=1, column=col+1, value=table_title[col])
    else:
        work_book = openpyxl.load_workbook(excel_path)
        work_sheet = work_book.active

    page = 1
    while True:
        # 开始爬取，一直到该分类结束
        root_url = root_url_prefix + str(page) + root_url_suffix
        try:
            root_html = requests.get(root_url, headers=header)
            root_html.encoding = 'utf-8'
            soup = BeautifulSoup(root_html.text, 'lxml')
        except:
            root_html = requests.get(root_url, headers=header)
            root_html.encoding = 'utf-8'
            soup = BeautifulSoup(root_html.text, 'lxml')
        paper_list = soup.select('.qt-info-title')
        if paper_list is None or len(paper_list) == 0:
            break
        paper_url_list = [paper.a.get('href') for paper in paper_list]
        for url in paper_url_list:
            url = 'http://5utk.ks5u.com/' + url
            rid = re.search(r'rid=(\d+)', url).group(1)
            try:
                html = requests.get(url, headers=header)
                html.encoding = 'utf-8'
                soup = BeautifulSoup(html.text, 'lxml')
            except:
                html = requests.get(url, headers=header)
                html.encoding = 'utf-8'
                soup = BeautifulSoup(html.text, 'lxml')
            if not os.path.exists(html_path + rid):
                os.mkdir(html_path + rid)
                with open(html_path + rid + '/' + rid + '.html', 'w') as f:
                    f.write(html.text)
            title = soup.select('title')[0].get_text()
            work_sheet.append([title, url, year, subject, grade])
            # split questions
            questions_list = soup.select('.bodyer_1')
            question_id = 1
            for question in questions_list:
                try:
                    problem = question.select('.bodyer_3')[0]
                    answer = question.select('.bodyer_5')[0]
                except IndexError:
                    # cannot find class bodyer_3 or bodyer_5
                    question_id += 1
                    continue
                with open(html_path + rid + '/' + rid + '_' + str(question_id) + '.html', 'w') as f:
                    f.write(problem.prettify())
                with open(html_path + rid + '/' + rid + '_' + str(question_id) + '_Answer.html', 'w') as f:
                    f.write(answer.prettify())
                question_id += 1

        page += 1

    # write xls
    work_book.save(filename=excel_path)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('-s', '--subject', type=str,
                        help='学科(chinese|math|english|physics|chemistry|biology|politics|history|geo)')
    parser.add_argument('-y', '--year', nargs='?', default='0', type=str,
                        help='年份(2021|2020|2019|2018|2017|2016|2015|2014|2013|2012|2011|2010)')
    parser.add_argument('-g', '--grade', nargs='?', default='0', type=str,
                        help='年级(s1|s2|s3)(senior high school)')
    parser.add_argument('-t', '--type', type=str,
                        help='试卷类型(高考真题|高考模拟|高中联考|期中试卷|期末试卷|月考试卷)')
    parser.add_argument('-ep', '--excel_path', type=str,
                        help='excel文件储存位置')
    parser.add_argument('-hp', '--html_path', type=str,
                        help='html文件储存位置')
    args = parser.parse_args()
    subject = args.subject
    year = args.year
    grade = args.grade
    exam_type = args.type
    excel_path = args.excel_path
    html_path = args.html_path
    run_spider(subject, year, grade, exam_type, excel_path, html_path)


if __name__ == '__main__':
    main()
